#!/usr/bin/env python3
"""
Chile Excel Vessel Data Processor
Handles Chile regional vessel registry Excel files with robust error handling
for format issues and complex Excel structures.
"""
import csv
import os
import sys
import warnings
from pathlib import Path
from typing import Dict, List, Optional, Any
import subprocess

# Suppress warnings for Excel processing
warnings.filterwarnings('ignore', category=UserWarning)

RAW_ROOT = Path(os.environ.get("EBISU_RAW_ROOT", "data/raw")).expanduser().resolve()
PROCESSED_ROOT = Path(os.environ.get("EBISU_PROCESSED_ROOT", RAW_ROOT)).expanduser().resolve()

def try_openpyxl_read(excel_file: Path) -> Optional[List[List[Any]]]:
    """Try to read Excel file with openpyxl (data_only mode)"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)

        # Get the first/main worksheet
        sheet = wb.active or wb[wb.sheetnames[0]]

        # Extract all data as list of lists
        data = []
        for row in sheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):  # Skip empty rows
                data.append(row)

        wb.close()
        return data

    except Exception as e:
        print(f"   ❌ openpyxl failed: {str(e)[:80]}...")
        return None

def try_pandas_read(excel_file: Path) -> Optional[List[List[Any]]]:
    """Try to read Excel file with pandas"""
    try:
        import pandas as pd

        # Try different engines
        for engine in ['openpyxl', 'xlrd', None]:
            try:
                df = pd.read_excel(excel_file, engine=engine, header=0)

                # Convert to list of lists (header + data)
                data = []
                data.append(df.columns.tolist())  # Header row
                data.extend(df.values.tolist())   # Data rows

                return data

            except Exception as engine_error:
                continue

    except Exception as e:
        print(f"   ❌ pandas failed: {str(e)[:80]}...")
        return None

def try_xlrd_direct(excel_file: Path) -> Optional[List[List[Any]]]:
    """Try to read Excel file with xlrd directly (for older .xls format)"""
    try:
        import xlrd

        workbook = xlrd.open_workbook(excel_file, ignore_workbook_corruption=True)
        sheet = workbook.sheet_by_index(0)

        data = []
        for row_idx in range(sheet.nrows):
            row_values = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                row_values.append(cell.value)
            data.append(row_values)

        return data

    except Exception as e:
        print(f"   ❌ xlrd failed: {str(e)[:80]}...")
        return None

def try_libreoffice_convert(excel_file: Path) -> Optional[List[List[Any]]]:
    """Try to convert Excel to CSV using LibreOffice and read that"""
    try:
        # Check if LibreOffice is available (try both soffice and libreoffice commands)
        soffice_cmd = None
        for cmd in ['soffice', 'libreoffice']:
            try:
                result = subprocess.run([cmd, '--version'],
                                      capture_output=True, text=True, timeout=5)
                if result.returncode == 0:
                    soffice_cmd = cmd
                    break
            except:
                continue

        if not soffice_cmd:
            return None

        # Create temp directory for conversion
        temp_dir = excel_file.parent / "temp_conversion"
        temp_dir.mkdir(exist_ok=True)

        # Convert to CSV
        csv_file = temp_dir / f"{excel_file.stem}.csv"
        convert_cmd = [
            soffice_cmd, '--headless', '--convert-to', 'csv',
            '--outdir', str(temp_dir), str(excel_file)
        ]

        result = subprocess.run(convert_cmd, capture_output=True, text=True, timeout=30)

        if csv_file.exists():
            # Read the converted CSV
            data = []
            with csv_file.open('r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    data.append(row)

            # Clean up temp file
            csv_file.unlink()
            temp_dir.rmdir()

            return data

    except Exception as e:
        print(f"   ❌ LibreOffice conversion failed: {str(e)[:80]}...")
        return None

def read_excel_robust(excel_file: Path) -> Optional[List[List[Any]]]:
    """Try multiple methods to read Excel file"""
    print(f"🔍 Reading: {excel_file.name}")

    # Try different reading methods in order of preference
    methods = [
        ("openpyxl data_only", try_openpyxl_read),
        ("pandas multi-engine", try_pandas_read),
        ("xlrd direct", try_xlrd_direct),
        ("LibreOffice convert", try_libreoffice_convert),
    ]

    for method_name, method_func in methods:
        print(f"   ⏳ Trying {method_name}...")
        data = method_func(excel_file)
        if data and len(data) > 1:  # Must have header + at least one data row
            print(f"   ✅ Success with {method_name}")
            return data

    print(f"   ❌ All methods failed for {excel_file.name}")
    return None

def clean_chile_data(data: List[List[Any]], region: str) -> List[Dict[str, str]]:
    """Clean and standardize Chile vessel data"""
    if not data or len(data) < 2:
        return []

    # Get headers from first row
    headers = data[0]

    # Clean headers: remove None, convert to string, standardize
    clean_headers = []
    for header in headers:
        if header is None:
            clean_headers.append("unknown_column")
        else:
            clean_name = str(header).strip().replace(' ', '_').replace('(', '').replace(')', '')
            clean_headers.append(clean_name)

    # Process data rows
    cleaned_rows = []
    for row_data in data[1:]:
        if not any(cell is not None and str(cell).strip() for cell in row_data):
            continue  # Skip empty rows

        row_dict = {}
        for i, value in enumerate(row_data):
            if i < len(clean_headers):
                # Clean the value
                clean_value = ""
                if value is not None:
                    clean_value = str(value).strip()
                    if clean_value.lower() in ['nan', 'none', '#n/a']:
                        clean_value = ""

                row_dict[clean_headers[i]] = clean_value

        # Add region identifier
        row_dict['chile_region'] = region
        cleaned_rows.append(row_dict)

    return cleaned_rows

def process_chile_excel(excel_file: Path) -> bool:
    """Process a single Chile Excel file"""
    try:
        # Extract region from filename: CHL_region_I_RPA_2025-09-08.xlsx -> I
        filename = excel_file.name
        if 'region_' in filename:
            region = filename.split('region_')[1].split('_')[0]
        else:
            region = excel_file.stem.replace('CHL_', '').replace('_2025-09-08', '')

        # Read Excel file
        raw_data = read_excel_robust(excel_file)
        if not raw_data:
            return False

        # Clean the data
        cleaned_data = clean_chile_data(raw_data, region)
        if not cleaned_data:
            print(f"   ⚠️ No valid data found in {excel_file.name}")
            return False

        # Output file
        out_dir = PROCESSED_ROOT / "vessels" / "COUNTRY" / "cleaned"
        out_dir.mkdir(parents=True, exist_ok=True)
        output_file = out_dir / f"chile_region_{region.lower()}_vessels_cleaned.csv"

        # Write cleaned CSV
        if cleaned_data:
            fieldnames = list(cleaned_data[0].keys())
            with output_file.open('w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(cleaned_data)

            print(f"✅ Processed {len(cleaned_data)} vessels → {output_file.name}")
            return True

        return False

    except Exception as e:
        print(f"❌ Failed to process {excel_file.name}: {e}")
        return False

def main():
    """Process all Chile Excel files"""
    chile_dir = RAW_ROOT / "vessels" / "vessel_data" / "COUNTRY" / "CHILE_LTP-PEP" / "raw"

    if not chile_dir.exists():
        print(f"❌ Chile directory not found: {chile_dir}")
        sys.exit(1)

    # Find Excel files
    excel_files = list(chile_dir.glob("*.xlsx")) + list(chile_dir.glob("*.xls"))
    if not excel_files:
        print(f"⚠️ No Excel files found in {chile_dir}")
        return

    print(f"🇨🇱 Processing {len(excel_files)} Chile Excel files...")

    successful = 0
    failed = 0

    for excel_file in sorted(excel_files):
        if process_chile_excel(excel_file):
            successful += 1
        else:
            failed += 1

    print(f"\n📊 Chile Excel Processing Complete:")
    print(f"   ✅ Successful: {successful}")
    print(f"   ❌ Failed: {failed}")
    print(f"   📁 Output: {PROCESSED_ROOT}/vessels/COUNTRY/cleaned/")

    if successful == 0:
        sys.exit(1)

if __name__ == "__main__":
    main()
