#!/usr/bin/env python3
"""
Simple converter for Chile Excel files that ignores styling issues
"""
import os
import sys
from pathlib import Path

# Try to use openpyxl directly
try:
    from openpyxl import load_workbook
    
    def convert_with_openpyxl(input_file, output_file):
        """Convert using openpyxl to avoid pandas formatting issues"""
        wb = load_workbook(input_file, data_only=True, read_only=True)
        ws = wb.active
        
        with open(output_file, 'w', encoding='utf-8') as f:
            # Write all rows
            for row in ws.iter_rows(values_only=True):
                # Clean and join values
                cleaned = []
                for val in row:
                    if val is None:
                        cleaned.append('')
                    else:
                        # Handle any type of value
                        cleaned.append(str(val).replace(',', ';').replace('\n', ' ').strip())
                f.write(','.join(cleaned) + '\n')
        
        wb.close()
        return True
        
except ImportError:
    print("openpyxl not available")
    sys.exit(1)

def main():
    source_dir = Path("/import/missing-registered-vessels-2025-09-08")
    
    # Process each Chile Excel file
    chile_files = list(source_dir.glob("CHL_*.xlsx"))
    
    if not chile_files:
        print("No Chile Excel files found")
        return 1
    
    print(f"Found {len(chile_files)} Chile Excel files to convert")
    
    successful = 0
    for excel_file in chile_files:
        # Extract region info from filename
        filename = excel_file.stem
        parts = filename.split('_')
        
        if 'region' in parts:
            region_idx = parts.index('region') + 1
            if region_idx < len(parts):
                region = parts[region_idx]
                dest_name = f"CHILE_{region}"
            else:
                dest_name = filename.replace('CHL_vessels_', 'CHILE_').split('_20')[0]
        else:
            dest_name = filename.replace('CHL_vessels_', 'CHILE_').split('_20')[0]
        
        # Create destination directory
        dest_dir = Path(f"/import/vessels/vessel_data/COUNTRY/{dest_name}/raw")
        dest_dir.mkdir(parents=True, exist_ok=True)
        
        # Define output CSV file
        csv_filename = f"{dest_name}_vessels_2025-09-08.csv"
        output_file = dest_dir / csv_filename
        
        # Convert
        try:
            if convert_with_openpyxl(excel_file, output_file):
                print(f"✓ Converted {excel_file.name} → {csv_filename}")
                successful += 1
        except Exception as e:
            print(f"✗ Error converting {excel_file.name}: {str(e)}")
    
    print(f"\nConversion complete: {successful}/{len(chile_files)} files converted successfully")
    return 0 if successful == len(chile_files) else 1

if __name__ == "__main__":
    sys.exit(main())